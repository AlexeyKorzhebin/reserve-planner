import datetime as dt
import math
from dataclasses import dataclass, field
from typing import Callable, Dict, Iterable, List, Optional, Tuple

import openpyxl


@dataclass
class DemandRow:
    rc: int
    rc_name: str
    plu: int
    plu_name: str
    date: dt.datetime
    qty: float


@dataclass
class ReserveInfo:
    supplier_name: Dict[int, str] = field(default_factory=dict)
    plu_name: Dict[int, str] = field(default_factory=dict)
    rc_name: Dict[int, str] = field(default_factory=dict)
    pallet_weight: Dict[int, float] = field(default_factory=dict)
    max_pallets: Dict[int, int] = field(default_factory=dict)
    reserve: Dict[Tuple[int, int], float] = field(default_factory=dict)
    suppliers_by_plu: Dict[int, List[int]] = field(default_factory=dict)


@dataclass
class TemplateEntry:
    pallets_by_supplier: Dict[int, int] = field(default_factory=dict)
    total_pallets: int = 0
    base_demand: int = 0


@dataclass
class TemplateData:
    entries: Dict[Tuple[int, int, dt.datetime], TemplateEntry] = field(default_factory=dict)
    shoulder_map: Dict[Tuple[int, int], int] = field(default_factory=dict)
    supplier_name: Dict[int, str] = field(default_factory=dict)
    plu_name: Dict[int, str] = field(default_factory=dict)
    rc_name: Dict[int, str] = field(default_factory=dict)
    pallet_weight: Dict[int, float] = field(default_factory=dict)


@dataclass
class AllocationItem:
    rc: int
    rc_name: str
    plu: int
    plu_name: str
    date: dt.datetime
    supplier: int
    supplier_name: str
    pallet_weight: float
    pallets: int


@dataclass
class OutputRow:
    date: dt.datetime
    supplier: int
    supplier_name: str
    plu: int
    plu_name: str
    rc: int
    rc_name: str
    volume: float
    truck_id: int
    shoulder: int
    pallet_weight: float
    pallets: int


@dataclass
class RunOptions:
    use_template: bool = True
    scale_template: bool = True
    include_template_without_demand: bool = True
    use_template_when_no_demand: bool = True
    ignore_reserve_limits: bool = True
    default_truck_pallets: int = 32


def _log(msg: str, logger: Optional[Callable[[str], None]]) -> None:
    if logger:
        logger(msg)


def _to_int(value, default: int = 0) -> int:
    if value is None:
        return default
    try:
        return int(value)
    except Exception:
        try:
            return int(float(str(value).replace(" ", "").replace(",", ".")))
        except Exception:
            return default


def _to_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except Exception:
        try:
            return float(str(value).replace(" ", "").replace(",", "."))
        except Exception:
            return default


def _to_date(value) -> Optional[dt.datetime]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return dt.datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _round_div(numer: int, denom: int) -> int:
    if denom <= 0:
        return 0
    return int((numer + denom / 2) // denom)


def load_demand(path: str, logger: Optional[Callable[[str], None]] = None) -> List[DemandRow]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx for idx, name in enumerate(header) if name}
    required = ["Код РЦ", "Наименование РЦ", "Код PLU", "Наименование PLU"]
    for col in required:
        if col not in col_map:
            raise ValueError(f"Не найдена колонка '{col}' в файле потребности.")
    date_cols = []
    for idx, name in enumerate(header):
        if idx <= col_map["Наименование PLU"]:
            continue
        if _to_date(name):
            date_cols.append(idx)
    if not date_cols:
        date_cols = list(range(col_map["Наименование PLU"] + 1, len(header)))
    rows: List[DemandRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rc = _to_int(row[col_map["Код РЦ"]])
        rc_name = row[col_map["Наименование РЦ"]] or ""
        plu = _to_int(row[col_map["Код PLU"]])
        plu_name = row[col_map["Наименование PLU"]] or ""
        for idx in date_cols:
            qty = _to_float(row[idx])
            if qty <= 0:
                continue
            date = _to_date(header[idx])
            if not date:
                continue
            rows.append(DemandRow(rc, rc_name, plu, plu_name, date, qty))
    wb.close()
    _log(f"Потребность: строк {len(rows)}", logger)
    return rows


def load_reserve(path: str, logger: Optional[Callable[[str], None]] = None) -> ReserveInfo:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx for idx, name in enumerate(header) if name}
    info = ReserveInfo()
    if "Код поставщика" not in col_map or "Код PLU" not in col_map:
        raise ValueError("Не найдены обязательные колонки в файле резерва.")
    idx_supplier_name = col_map.get("Наименование поставщика")
    idx_plu_name = col_map.get("Наименование PLU")
    idx_reserve = col_map.get("Резерв")
    idx_max_pallets = col_map.get("Максимальное кол-во паллет в машине (макс. квант), шт.")
    idx_min_pallets = col_map.get("Минимальное кол-во паллет в машине (мин. квант), шт.")
    idx_pallet_weight = col_map.get("Макс. вес товара на одной паллете, кг")
    for row in ws.iter_rows(min_row=2, values_only=True):
        supplier = _to_int(row[col_map["Код поставщика"]])
        supplier_name = row[idx_supplier_name] if idx_supplier_name is not None else ""
        supplier_name = supplier_name or ""
        plu = _to_int(row[col_map["Код PLU"]])
        plu_name = row[idx_plu_name] if idx_plu_name is not None else ""
        plu_name = plu_name or ""
        reserve_val = _to_float(row[idx_reserve]) if idx_reserve is not None else 0.0
        max_pallets = _to_int(row[idx_max_pallets]) if idx_max_pallets is not None else 0
        if not max_pallets:
            max_pallets = _to_int(row[idx_min_pallets]) if idx_min_pallets is not None else 0
        pallet_weight = _to_float(row[idx_pallet_weight]) if idx_pallet_weight is not None else 0.0

        if supplier:
            if supplier_name:
                info.supplier_name[supplier] = supplier_name
            if max_pallets:
                info.max_pallets[supplier] = max_pallets
        if plu:
            if plu_name:
                info.plu_name[plu] = plu_name
            if pallet_weight:
                info.pallet_weight[plu] = pallet_weight
        if supplier and plu and reserve_val:
            info.reserve[(supplier, plu)] = reserve_val
            info.suppliers_by_plu.setdefault(plu, []).append(supplier)
    wb.close()
    _log(f"Резерв: поставщиков {len(info.supplier_name)}, PLU {len(info.plu_name)}", logger)
    return info


def load_template(path: str, logger: Optional[Callable[[str], None]] = None) -> TemplateData:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx for idx, name in enumerate(header) if name}
    required = [
        "Дата поставки",
        "SAP поставщика",
        "Поставщик",
        "PLU",
        "Наименование ",
        "РЦ",
        "Наименование РЦ",
        "Объем",
        "лог плечо",
        "Квант паллет",
        "паллеты",
    ]
    for col in required:
        if col not in col_map:
            raise ValueError(f"Не найдена колонка '{col}' в файле образца.")

    data = TemplateData()
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = _to_date(row[col_map["Дата поставки"]])
        if not date:
            continue
        supplier = _to_int(row[col_map["SAP поставщика"]])
        supplier_name = row[col_map["Поставщик"]] or ""
        plu = _to_int(row[col_map["PLU"]])
        plu_name = row[col_map["Наименование "]] or ""
        rc = _to_int(row[col_map["РЦ"]])
        rc_name = row[col_map["Наименование РЦ"]] or ""
        volume = _to_float(row[col_map["Объем"]])
        shoulder = _to_int(row[col_map["лог плечо"]])
        pallet_weight = _to_float(row[col_map["Квант паллет"]])
        pallets = _to_int(row[col_map["паллеты"]])
        if not pallets and pallet_weight:
            pallets = _round_div(int(round(volume)), int(round(pallet_weight)))
        key = (rc, plu, date)
        entry = data.entries.setdefault(key, TemplateEntry())
        entry.pallets_by_supplier[supplier] = entry.pallets_by_supplier.get(supplier, 0) + pallets
        entry.total_pallets += pallets

        if supplier_name:
            data.supplier_name[supplier] = supplier_name
        if rc_name:
            data.rc_name[rc] = rc_name
        if plu_name:
            data.plu_name[plu] = plu_name
        if pallet_weight:
            data.pallet_weight[plu] = pallet_weight
        if supplier and rc and shoulder:
            data.shoulder_map.setdefault((supplier, rc), shoulder)
    wb.close()
    _log(f"Образец: ключей {len(data.entries)}", logger)
    return data


def _scale_pallets(
    pallets_by_supplier: Dict[int, int], base_demand: int, demand_value: float
) -> Dict[int, int]:
    demand_int = max(0, int(round(demand_value)))
    if base_demand <= 0 or demand_int <= 0:
        return {}
    total_pallets = sum(pallets_by_supplier.values())
    desired_total = _round_div(total_pallets * demand_int, base_demand)

    scaled: Dict[int, int] = {}
    remainders: List[Tuple[int, int]] = []
    total_scaled = 0
    for supplier, pallets in pallets_by_supplier.items():
        numer = pallets * demand_int
        base = numer // base_demand
        rem = numer % base_demand
        scaled[supplier] = int(base)
        total_scaled += int(base)
        remainders.append((supplier, rem))

    delta = desired_total - total_scaled
    if delta > 0:
        remainders.sort(key=lambda x: x[1], reverse=True)
        for supplier, _ in remainders[:delta]:
            scaled[supplier] += 1
    elif delta < 0:
        remainders.sort(key=lambda x: x[1])
        for supplier, _ in remainders[: abs(delta)]:
            scaled[supplier] = max(0, scaled[supplier] - 1)
    return {k: v for k, v in scaled.items() if v > 0}


def build_allocations(
    demand_rows: Iterable[DemandRow],
    reserve: ReserveInfo,
    template: Optional[TemplateData],
    options: RunOptions,
    logger: Optional[Callable[[str], None]] = None,
) -> List[AllocationItem]:
    demand_by_key: Dict[Tuple[int, int, dt.datetime], float] = {}
    for row in demand_rows:
        key = (row.rc, row.plu, row.date)
        demand_by_key[key] = demand_by_key.get(key, 0.0) + row.qty
        if row.rc_name:
            reserve.rc_name.setdefault(row.rc, row.rc_name)
        if row.plu_name:
            reserve.plu_name.setdefault(row.plu, row.plu_name)

    if template:
        for key, entry in template.entries.items():
            entry.base_demand = int(round(demand_by_key.get(key, 0.0)))

    items: List[AllocationItem] = []
    for row in demand_rows:
        key = (row.rc, row.plu, row.date)
        pallet_weight = reserve.pallet_weight.get(row.plu) or (
            template.pallet_weight.get(row.plu) if template else None
        )
        if not pallet_weight:
            pallet_weight = 1.0
        supplier_name = ""
        if template and options.use_template and key in template.entries:
            entry = template.entries[key]
            if options.scale_template and entry.base_demand > 0:
                pallets_by_supplier = _scale_pallets(
                    entry.pallets_by_supplier, entry.base_demand, row.qty
                )
            elif options.use_template_when_no_demand or not options.scale_template:
                pallets_by_supplier = dict(entry.pallets_by_supplier)
            else:
                pallets_by_supplier = {}
        else:
            pallets_total = int(math.ceil(row.qty / pallet_weight))
            suppliers = reserve.suppliers_by_plu.get(row.plu, [])
            supplier = suppliers[0] if suppliers else 0
            pallets_by_supplier = {supplier: pallets_total} if supplier else {}

        for supplier, pallets in pallets_by_supplier.items():
            if pallets <= 0:
                continue
            supplier_name = reserve.supplier_name.get(supplier) or (
                template.supplier_name.get(supplier) if template else ""
            )
            items.append(
                AllocationItem(
                    rc=row.rc,
                    rc_name=reserve.rc_name.get(row.rc, row.rc_name),
                    plu=row.plu,
                    plu_name=reserve.plu_name.get(row.plu, row.plu_name),
                    date=row.date,
                    supplier=supplier,
                    supplier_name=supplier_name,
                    pallet_weight=pallet_weight,
                    pallets=pallets,
                )
            )

    if template and options.use_template and options.include_template_without_demand:
        for key, entry in template.entries.items():
            if key in demand_by_key:
                continue
            rc, plu, date = key
            pallet_weight = reserve.pallet_weight.get(plu) or template.pallet_weight.get(plu) or 1.0
            for supplier, pallets in entry.pallets_by_supplier.items():
                if pallets <= 0:
                    continue
                supplier_name = reserve.supplier_name.get(supplier) or template.supplier_name.get(supplier, "")
                items.append(
                    AllocationItem(
                        rc=rc,
                        rc_name=reserve.rc_name.get(rc, template.rc_name.get(rc, "")),
                        plu=plu,
                        plu_name=reserve.plu_name.get(plu, template.plu_name.get(plu, "")),
                        date=date,
                        supplier=supplier,
                        supplier_name=supplier_name,
                        pallet_weight=pallet_weight,
                        pallets=pallets,
                    )
                )

    _log(f"Аллокации: строк {len(items)}", logger)
    return items


def pack_trucks(
    items: List[AllocationItem],
    reserve: ReserveInfo,
    template: Optional[TemplateData],
    options: RunOptions,
    logger: Optional[Callable[[str], None]] = None,
) -> List[OutputRow]:
    grouped: Dict[int, List[AllocationItem]] = {}
    for item in items:
        grouped.setdefault(item.supplier, []).append(item)

    rows: List[OutputRow] = []
    truck_id = 1
    for supplier in sorted(grouped.keys()):
        cap = reserve.max_pallets.get(supplier, options.default_truck_pallets)
        remaining = 0
        current_truck = 0
        items_sorted = sorted(
            grouped[supplier],
            key=lambda x: (x.date, x.rc, x.plu),
        )
        for item in items_sorted:
            pallets_left = item.pallets
            while pallets_left > 0:
                if remaining == 0:
                    current_truck = truck_id
                    truck_id += 1
                    remaining = cap
                take = min(remaining, pallets_left)
                remaining -= take
                pallets_left -= take
                shoulder = 1
                if template:
                    shoulder = template.shoulder_map.get((supplier, item.rc), 1)
                rows.append(
                    OutputRow(
                        date=item.date,
                        supplier=supplier,
                        supplier_name=item.supplier_name,
                        plu=item.plu,
                        plu_name=item.plu_name,
                        rc=item.rc,
                        rc_name=item.rc_name,
                        volume=take * item.pallet_weight,
                        truck_id=current_truck,
                        shoulder=shoulder,
                        pallet_weight=item.pallet_weight,
                        pallets=take,
                    )
                )
    _log(f"Выходные строки: {len(rows)}", logger)
    return rows


def write_output(path: str, rows: List[OutputRow], logger: Optional[Callable[[str], None]] = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Распределение"
    headers = [
        "Дата поставки",
        "SAP поставщика",
        "Поставщик",
        "PLU",
        "Наименование ",
        "РЦ",
        "Наименование РЦ",
        "Объем",
        "ТС",
        "лог плечо",
        "Квант паллет",
        "паллеты",
        "тс",
    ]
    ws.append(headers)
    for row in sorted(rows, key=lambda x: (x.date, x.supplier, x.truck_id, x.rc, x.plu)):
        ws.append(
            [
                row.date,
                row.supplier,
                row.supplier_name,
                row.plu,
                row.plu_name,
                row.rc,
                row.rc_name,
                row.volume,
                row.truck_id,
                row.shoulder,
                row.pallet_weight,
                row.pallets,
                None,
            ]
        )
    wb.save(path)
    _log(f"Сохранено: {path}", logger)


def run_plan(
    demand_path: str,
    reserve_path: str,
    output_path: str,
    template_path: Optional[str] = None,
    options: Optional[RunOptions] = None,
    logger: Optional[Callable[[str], None]] = None,
) -> None:
    opts = options or RunOptions()
    demand_rows = load_demand(demand_path, logger)
    reserve = load_reserve(reserve_path, logger)
    template = load_template(template_path, logger) if (template_path and opts.use_template) else None
    allocations = build_allocations(demand_rows, reserve, template, opts, logger)
    output_rows = pack_trucks(allocations, reserve, template, opts, logger)
    write_output(output_path, output_rows, logger)
